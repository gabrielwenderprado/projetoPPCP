from pathlib import Path
import json
from openpyxl import load_workbook

xlsx = Path('/home/ubuntu/upload/CópiadeEXPLOSAO_12.08(2).xlsm')
wb = load_workbook(xlsx, data_only=True, read_only=True)
ws = wb['Programacao']
print('sheet=', ws.title, 'max_row=', ws.max_row, 'max_column=', ws.max_column)
for row_number in range(1, 8):
    values = [ws.cell(row_number, col).value for col in range(1, min(ws.max_column, 53) + 1)]
    print('row', row_number, 'A=', values[0] if values else None, 'AS=', values[44] if len(values) >= 45 else None)
    if row_number <= 3:
        print('headers=', [(index + 1, value) for index, value in enumerate(values) if value not in (None, '')])

data = json.loads(Path('/home/ubuntu/PCM-ATT_9.0/data/explosao.json').read_text(encoding='utf-8'))
print('json_item_keys=', sorted(data['items'][0].keys()))
print('json_items=', len(data['items']))
