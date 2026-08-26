from pathlib import Path
import json
import re
import sys

from openpyxl import load_workbook


HEADER_ROW = 3
MODEL_ROW = 2
FIRST_DATA_ROW = 4
OUTPUT_DEFAULT = Path(__file__).resolve().parents[1] / "data" / "pinos.json"


def text(value):
    return "" if value is None else str(value).strip()


def number(value):
    if isinstance(value, (int, float)):
        return float(value or 0)
    raw = text(value).replace(".", "").replace(",", ".")
    try:
        return float(raw) if raw else 0
    except ValueError:
        return 0


def compact(value):
    value = float(value or 0)
    return int(value) if value.is_integer() else round(value, 4)


def unique_name(name, used):
    base = re.sub(r"^PINOS\s*", "", text(name), flags=re.I).strip() or "Modelo"
    result = base
    index = 2
    while result in used:
        result = f"{base} ({index})"
        index += 1
    return result


def extrair(origem: Path, destino: Path) -> None:
    wb = load_workbook(origem, read_only=False, data_only=True, keep_vba=True)
    sheet_name = next((name for name in ("Pinos", "pinos") if name in wb.sheetnames), None)
    if not sheet_name:
        raise ValueError("A aba Pinos não foi encontrada na planilha.")
    ws = wb[sheet_name]

    code_columns = []
    used_models = set()
    for col in range(1, ws.max_column + 1):
        if text(ws.cell(HEADER_ROW, col).value).lower() != "código":
            continue
        model = ""
        for candidate_col in range(col, min(ws.max_column, col + 8) + 1):
            candidate = text(ws.cell(MODEL_ROW, candidate_col).value)
            if candidate:
                model = unique_name(candidate, used_models)
                break
        if not model:
            model = unique_name(f"Modelo {len(code_columns) + 1}", used_models)
        used_models.add(model)
        code_columns.append({
            "model": model,
            "code": col,
            "description": col + 1,
            "stock": col + 3,
            "need": col + 4,
        })

    items = {}
    model_order = [block["model"] for block in code_columns]
    for block in code_columns:
        model = block["model"]
        for row in range(FIRST_DATA_ROW, ws.max_row + 1):
            code = text(ws.cell(row, block["code"]).value)
            if not code or code.lower() in {"código", "codigo"}:
                continue
            description = text(ws.cell(row, block["description"]).value)
            item = items.setdefault(code, {
                "code": code,
                "description": description,
                "unit": text(ws.cell(row, block["description"] + 1).value) or "UN",
                "stock": 0,
                "modelNeeds": {name: 0 for name in model_order},
            })
            if not item["description"] and description:
                item["description"] = description
            stock = compact(number(ws.cell(row, block["stock"]).value))
            need = compact(number(ws.cell(row, block["need"]).value))
            if item["stock"] == 0 and stock:
                item["stock"] = stock
            item["modelNeeds"][model] = compact(number(item["modelNeeds"].get(model, 0)) + need)

    output_items = []
    for item in items.values():
        item["totalNeed"] = compact(sum(number(value) for value in item["modelNeeds"].values()))
        item["modelCount"] = sum(1 for value in item["modelNeeds"].values() if number(value) > 0)
        output_items.append(item)
    output_items.sort(key=lambda item: (-(number(item["totalNeed"])), item["code"]))

    payload = {
        "sourceFile": origem.name,
        "sourceSheet": sheet_name,
        "generatedAt": __import__("datetime").date.today().isoformat(),
        "models": model_order,
        "items": output_items,
        "meta": {"headerRow": HEADER_ROW, "firstDataRow": FIRST_DATA_ROW, "codeBlocks": len(code_columns)},
    }
    destino.parent.mkdir(parents=True, exist_ok=True)
    destino.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"PINOS_OK {len(output_items)} itens | {len(model_order)} modelos | {destino}")


if __name__ == "__main__":
    if len(sys.argv) not in {2, 3}:
        raise SystemExit("Uso: python convert_pinos.py origem.xlsm [destino.json]")
    extrair(Path(sys.argv[1]), Path(sys.argv[2]) if len(sys.argv) == 3 else OUTPUT_DEFAULT)
