#!/usr/bin/env python3
"""Executa e verifica a atualização completa dos snapshots do dashboard.

Uso:
  python scripts/atualizar_todos_dados.py explosao.xlsm [consumiveis.xlsx]

A planilha de Consumíveis é opcional porque, no projeto original, ela é uma
origem separada da planilha de Explosão. Se a aba consumiveis existir na
primeira planilha, ela será usada automaticamente.
"""
from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
DATA = ROOT / "data"


def fail(message: str) -> None:
    raise SystemExit(f"ERRO: {message}")


def has_sheet(source: Path, expected: str) -> bool:
    try:
        wb = load_workbook(source, read_only=True, data_only=True)
    except TypeError:
        wb = load_workbook(source, read_only=True, data_only=True)
    try:
        return any(name.strip().lower() == expected.lower() for name in wb.sheetnames)
    finally:
        wb.close()


def run(command: list[str]) -> None:
    print(f"\n> {' '.join(command)}")
    result = subprocess.run(command, cwd=ROOT)
    if result.returncode != 0:
        fail(f"o conversor terminou com código {result.returncode}")


def verify_json(filename: str, required: tuple[str, ...]) -> dict:
    path = DATA / filename
    if not path.exists():
        fail(f"snapshot não foi gerado: {path}")
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        fail(f"snapshot inválido {filename}: {exc}")
    if not all(key in payload for key in required):
        fail(f"snapshot {filename} não contém o schema esperado: {required}")
    return payload


def main() -> int:
    if len(sys.argv) < 2:
        fail("informe a planilha de Explosão; o Consumível pode ser informado como segundo argumento")
    explosion = Path(sys.argv[1]).expanduser().resolve()
    if not explosion.exists():
        fail(f"planilha de Explosão não encontrada: {explosion}")
    consumables = Path(sys.argv[2]).expanduser().resolve() if len(sys.argv) >= 3 else None
    if consumables and not consumables.exists():
        fail(f"planilha de Consumíveis não encontrada: {consumables}")

    run([sys.executable, str(SCRIPTS / "convert_excel.py"), str(explosion)])

    embedded_consumables = has_sheet(explosion, "consumiveis")
    if consumables is None and embedded_consumables:
        consumables = explosion
    if consumables is not None:
        run([sys.executable, str(SCRIPTS / "convert_consumiveis.py"), str(consumables)])
    else:
        print("AVISO: Consumíveis não atualizado: informe a segunda planilha ou use uma origem com a aba consumiveis.")

    snapshots = {
        "explosao.json": ("sourceFile", "generatedAt", "items", "models"),
        "plano-mes.json": ("sourceSheet", "months", "models"),
        "pinos.json": ("sourceFile", "sourceSheet", "models", "items"),
        "cilindros.json": ("sourceFile", "sourceSheet", "models", "items"),
        "historico-estoque.json": ("description", "records"),
    }
    if consumables is not None:
        snapshots["consumiveis.json"] = ("sourceFile", "sheet", "months", "items")

    print("\nRESUMO DA ATUALIZAÇÃO")
    for filename, required in snapshots.items():
        payload = verify_json(filename, required)
        sizes = []
        if isinstance(payload.get("items"), list):
            sizes.append(f"itens={len(payload['items'])}")
        if isinstance(payload.get("models"), list):
            sizes.append(f"modelos={len(payload['models'])}")
        if isinstance(payload.get("records"), list):
            sizes.append(f"registos={len(payload['records'])}")
        print(f"OK {filename}: {', '.join(sizes) or 'schema válido'}")
    print("\nATUALIZAÇÃO_COMPLETA_OK")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
