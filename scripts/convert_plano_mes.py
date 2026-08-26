"""Extrai as linhas 15 a 27 da aba PLANO MES para o snapshot do dashboard."""
from pathlib import Path
import json
import sys

from openpyxl import load_workbook


MESES_PADRAO = ["06", "07", "08", "09", "10", "11", "12"]


def valor(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def extrair(origem: Path, destino: Path) -> None:
    wb = load_workbook(origem, read_only=True, data_only=True, keep_vba=True)
    if "PLANO MES" not in wb.sheetnames:
        raise ValueError("A aba PLANO MES não foi encontrada na planilha.")
    ws = wb["PLANO MES"]
    cabecalhos = [valor(ws.cell(15, col).value) for col in range(1, ws.max_column + 1)]
    mes_por_coluna = {}
    for col, cabecalho in enumerate(cabecalhos, start=1):
        if cabecalho is None:
            continue
        texto = str(cabecalho).strip()
        if texto.isdigit() and len(texto) <= 2:
            mes_por_coluna[col] = texto.zfill(2)
    meses = []
    for mes in MESES_PADRAO:
        if mes in mes_por_coluna.values():
            meses.append(mes)
    for mes in mes_por_coluna.values():
        if mes not in meses:
            meses.append(mes)

    modelos = []
    for linha in range(16, 31):
        modelo = valor(ws.cell(linha, 1).value)
        if modelo is None or str(modelo).strip() == "":
            continue
        quantidades = {}
        for col, mes in mes_por_coluna.items():
            quantidades[mes] = valor(ws.cell(linha, col).value) or 0
        modelos.append({"modelo": str(modelo).strip(), "quantidades": quantidades})

    payload = {
        "sourceSheet": "PLANO MES",
        "sourceRange": "A15:Y30",
        "months": meses,
        "models": modelos,
    }
    destino.parent.mkdir(parents=True, exist_ok=True)
    destino.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        raise SystemExit("Uso: python convert_plano_mes.py origem.xlsm destino.json")
    extrair(Path(sys.argv[1]), Path(sys.argv[2]))
    print("PLANO_MES_OK")
