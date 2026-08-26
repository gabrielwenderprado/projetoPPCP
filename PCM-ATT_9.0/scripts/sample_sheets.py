from openpyxl import load_workbook
source='/home/ubuntu/upload/CópiadeEXPLOSAO_04.08.xlsm'
wb=load_workbook(source, read_only=True, data_only=True, keep_vba=True)
for name in ['Estoque','Obtencao','Programacao','familia','estoque segurança','10S','13ldi-46kv','13-69kv','13AT']:
    ws=wb[name]
    print(f'\n### {name} ({ws.max_row}x{ws.max_column})')
    for row in ws.iter_rows(min_row=1, max_row=min(8, ws.max_row), max_col=min(20, ws.max_column), values_only=True):
        print([str(v)[:70] if v is not None else '' for v in row])
