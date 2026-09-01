# Biblioteca usada para trabalhar com caminhos de ficheiros de forma compatível entre sistemas.
from pathlib import Path
# Biblioteca que lê ficheiros Excel, inclusive planilhas com macros em modo de leitura.
from openpyxl import load_workbook
import json
import re
import sys
import unicodedata
from datetime import date, datetime

from convert_plano_mes import extrair as extrair_plano_mes
from convert_pinos import extrair as extrair_pinos
from convert_cilindros import extrair as extrair_cilindros
from convert_cabines import extrair as extrair_cabines

# Define a raiz do projeto a partir da localização deste script.
PROJECT_ROOT = Path(__file__).resolve().parents[1]
# Usa o ficheiro informado no comando ou um caminho padrão de exemplo.
SOURCE = Path(sys.argv[1]) if len(sys.argv) > 1 else Path('/home/ubuntu/upload/CópiadeEXPLOSAO_04.08.xlsm')
# Define onde os JSONs atualizados serão gravados.
OUTPUT = PROJECT_ROOT / 'data' / 'explosao.json'
PLANO_OUTPUT = PROJECT_ROOT / 'data' / 'plano-mes.json'
PINOS_OUTPUT = PROJECT_ROOT / 'data' / 'pinos.json'
CILINDROS_OUTPUT = PROJECT_ROOT / 'data' / 'cilindros.json'
CABINES_OUTPUT = PROJECT_ROOT / 'data' / 'cabines.json'
HISTORY_OUTPUT = PROJECT_ROOT / 'data' / 'historico-estoque.json'
MODEL_SHEETS = [
    '10S', '13ldi-46kv', '13-69kv', '13AT', '15LDDI', '18lddi',
    'guin-16T', 'guin-25T', 'guin-45', 'replica-helio-base', '10L', '10HDOC', '8paletes', '30-T','21T','12-T','7T','basculante'
]


def text(value):
    return '' if value is None else str(value).strip()


def normalize(value):
    value = unicodedata.normalize('NFKD', text(value))
    value = ''.join(char for char in value if not unicodedata.combining(char))
    return re.sub(r'\s+', ' ', value).strip().lower()


def num(value):
    if isinstance(value, (int, float)):
        return float(value or 0)
    raw = text(value).replace('.', '').replace(',', '.')
    try:
        return float(raw) if raw else 0
    except ValueError:
        return 0


def compact_number(value):
    value = float(value or 0)
    return int(value) if value.is_integer() else round(value, 4)


def rows(ws, max_row, max_col=None):
    kwargs = {'min_row': 1, 'max_row': max_row}
    if max_col:
        kwargs['max_col'] = max_col
    for row in ws.iter_rows(**kwargs, values_only=True):
        yield list(row)


# Abre a planilha somente para leitura e mantém a compatibilidade com ficheiros XLSM.
wb = load_workbook(SOURCE, read_only=True, data_only=True, keep_vba=True)
required = ['Estoque', 'Obtencao', 'Programacao', 'familia', 'estoque segurança']
missing = [sheet for sheet in required if sheet not in wb.sheetnames]
if missing:
    raise RuntimeError('Abas obrigatórias em falta: ' + ', '.join(missing))

stock = {}
ws = wb['Estoque']
for row in rows(ws, min(ws.max_row, 40000), 6):
    if len(row) >= 5 and text(row[1]) and text(row[1]) != 'Código Item Ref.':
        stock[text(row[1])] = {
            'description': text(row[2]),
            'unit': text(row[3]),
            'stock': compact_number(num(row[4])),
        }

safety = {}
ws = wb['estoque segurança']
for row in rows(ws, min(ws.max_row, 10000), 2):
    if len(row) >= 2 and text(row[0]) and normalize(row[0]) != 'codigo':
        safety[text(row[0])] = compact_number(num(row[1]))

ws = wb['Programacao']
header = next(rows(ws, 1, 60))
headers = [text(value) for value in header]


def idx_any(*names):
    wanted = {normalize(name) for name in names}
    return next((index for index, value in enumerate(headers) if normalize(value) in wanted), -1)


code_i = idx_any('Código Item')
desc_i = idx_any('Descrição Item')
analyst_i = idx_any('analista')
stock_i = idx_any(' estoque', 'estoque')
stock_max_i = idx_any('estoque maximo', 'estoque máximo', 'estoque max', 'estoque máximo')
stock_value_i = idx_any('valor de estoque', 'valor estoque')
safety_i = idx_any('estoq.segurança', 'estoque segurança')
family_i = idx_any('familia', 'família')
obtention_i = idx_any('Tipo de obtenção', 'Tipo obtenção', 'Tipo de obtencao')
movement_i = idx_any('movimentação', 'movimentacao')
if analyst_i < 0:
    raise RuntimeError('A coluna analista é obrigatória na aba Programacao')

order_cols = []
demand_cols = []
for index, header_value in enumerate(headers):
    normalized = re.sub(r'\s+', ' ', header_value.upper()).strip()
    order = re.search(r'PED\s+(\d{2}(?:/\d{4})?)', normalized)
    if order:
        order_cols.append((index, f'PED {order.group(1)}'))
    demand = re.search(r'DEM(?:ANDA)?\s+(\d{2}/\d{4})', normalized)
    if demand:
        demand_cols.append((index, f'DEM {demand.group(1)}'))

months = []
for _, label in order_cols:
    if label not in months:
        months.append(label)
demand_months = []
for _, label in demand_cols:
    if label not in demand_months:
        demand_months.append(label)

sc_cols = []
for index, header_value in enumerate(headers):
    if normalize(header_value) == 'colocar sc':
        previous = next((label for item_index, label in reversed(demand_cols) if item_index < index), '')
        sc_cols.append((index, previous))

items = []
seen = set()
analysts = set()
families = set()
obtention_types = set()
for row in rows(ws, min(ws.max_row, 30000), 60):
    code = text(row[code_i]) if 0 <= code_i < len(row) else ''
    if not code or normalize(code) == 'codigo item' or code.startswith('='):
        continue
    desc = text(row[desc_i]) if 0 <= desc_i < len(row) else stock.get(code, {}).get('description', '')
    analyst = text(row[analyst_i]) if 0 <= analyst_i < len(row) else ''
    if normalize(analyst) in {'analista', 'responsavel'}:
        analyst = ''
    family = text(row[family_i]) if 0 <= family_i < len(row) else ''
    obtention_type = text(row[obtention_i]) if 0 <= obtention_i < len(row) else ''
    movement_value = row[movement_i] if 0 <= movement_i < len(row) else None
    if isinstance(movement_value, (datetime, date)):
        last_movement = movement_value.strftime('%Y-%m-%d')
    else:
        last_movement = text(movement_value) or 'nao tem'
    orders = {label: compact_number(num(row[index])) for index, label in order_cols if index < len(row)}
    demands = {label: compact_number(num(row[index])) for index, label in demand_cols if index < len(row)}
    colocar_sc = {label: compact_number(num(row[index])) for index, label in sc_cols if label and index < len(row)}
    stock_value = compact_number(num(row[stock_i])) if 0 <= stock_i < len(row) else stock.get(code, {}).get('stock', 0)
    item = {
        'code': code,
        'description': desc,
        'analyst': analyst,
        'family': family,
        'obtentionType': obtention_type,
        'lastMovement': last_movement,
        'unit': stock.get(code, {}).get('unit', 'UN'),
        'stock': stock_value,
        'stockMax': compact_number(num(row[stock_max_i])) if 0 <= stock_max_i < len(row) else 0,
        'stockValue': compact_number(num(row[stock_value_i])) if 0 <= stock_value_i < len(row) else 0,
        'safety': compact_number(num(row[safety_i])) if 0 <= safety_i < len(row) else safety.get(code, 0),
        'orders': orders,
        'demands': demands,
        'colocarSC': colocar_sc,
    }
    if code not in seen:
        items.append(item)
        seen.add(code)
    if analyst:
        analysts.add(analyst)
    if family:
        families.add(family)
    if obtention_type:
        obtention_types.add(obtention_type)

models = {}
for name in MODEL_SHEETS:
    if name not in wb.sheetnames:
        continue
    ws = wb[name]
    sheet_rows = list(rows(ws, min(ws.max_row, 6000), min(ws.max_column, 12)))
    if not sheet_rows:
        continue
    model_header = [normalize(value) for value in sheet_rows[0]]
    code_index = next((index for index, value in enumerate(model_header) if value in {'codigo', 'codigo item'}), 1)
    description_index = next((index for index, value in enumerate(model_header) if value.startswith('descricao')), code_index + 1)
    quantity_index = next((index for index, value in enumerate(model_header) if value.startswith('quantidade')), code_index + 2)
    model_items = []
    seen_model_codes = set()
    for row in sheet_rows[1:]:
        code = text(row[code_index]) if code_index < len(row) else ''
        if not code or normalize(code) in {'codigo', 'codigo item'} or code.startswith('='):
            continue
        if code in seen_model_codes:
            continue
        model_items.append({
            'code': code,
            'description': text(row[description_index]) if description_index < len(row) else '',
            'quantity': compact_number(num(row[quantity_index])) if quantity_index < len(row) else 0,
        })
        seen_model_codes.add(code)
    if model_items:
        models[name] = model_items

open_requests = 0
open_quantity = 0
ws = wb['Obtencao']
for row in rows(ws, min(ws.max_row, 10000), 20):
    if len(row) > 5 and text(row[0]) and text(row[0]) != 'Nro. Solicitação' and normalize(row[5]) not in {'fechado', 'encerrado', 'cancelado'}:
        open_requests += 1
        open_quantity += num(row[10]) if len(row) > 10 else 0

# Reúne os dados normalizados no formato consumido pelo dashboard.
today = date.today()
month_names = {
    1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
    5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
    9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro',
}
payload = {
    'sourceFile': SOURCE.name,
    'generatedAt': today.isoformat(),
    'items': items,
    'models': models,
    'analysts': sorted(analysts),
    'families': sorted(families),
    'obtentionTypes': sorted(obtention_types),
    'months': months,
    'demandMonths': demand_months,
    'openRequests': open_requests,
    'openQuantity': compact_number(open_quantity),
    'meta': {
        'sheets': wb.sheetnames,
        'requiredSheets': required,
        'limits': {'programacaoRows': 30000, 'modelRows': 6000},
    },
}
OUTPUT.parent.mkdir(parents=True, exist_ok=True)
OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
# Mantém o Plano Mês sincronizado no mesmo arraste do Excel.
extrair_plano_mes(SOURCE, PLANO_OUTPUT)
# Mantém os pinos sincronizados com a mesma explosão Excel.
extrair_pinos(SOURCE, PINOS_OUTPUT)
# Mantém os cilindros sincronizados com a aba cilindros da mesma explosão Excel.
extrair_cilindros(SOURCE, CILINDROS_OUTPUT)
# Mantém as cabines sincronizadas com a aba cabines da mesma explosão Excel.
extrair_cabines(SOURCE, CABINES_OUTPUT)

# Acrescenta ou substitui o registo do mês atual com os números da nova planilha.
try:
    history = json.loads(HISTORY_OUTPUT.read_text(encoding='utf-8')) if HISTORY_OUTPUT.exists() else {'records': []}
except json.JSONDecodeError:
    history = {'records': []}
record = {
    'date': today.isoformat(),
    'label': f'{month_names[today.month]} {today.year}',
    'stockValue': round(sum(num(item.get('stockValue')) for item in items), 2),
    'itemCount': len(items),
    'sourceFile': SOURCE.name,
}
records = [item for item in history.get('records', []) if not str(item.get('date', '')).startswith(today.strftime('%Y-%m'))]
records.append(record)
records.sort(key=lambda item: str(item.get('date', '')))
HISTORY_OUTPUT.write_text(json.dumps({'description': 'Histórico mensal do valor total do estoque e da quantidade de materiais.', 'records': records}, ensure_ascii=False, indent=2), encoding='utf-8')
print(f'Histórico atualizado: {record["label"]} | {record["stockValue"]:.2f} | {record["itemCount"]} itens | {HISTORY_OUTPUT}')
print(f'Plano Mês atualizado: {PLANO_OUTPUT}')
print(f'Pinos atualizados: {PINOS_OUTPUT}')
print(f'Cilindros atualizados: {CILINDROS_OUTPUT}')
print(f'Cabines atualizadas: {CABINES_OUTPUT}')
print(f'OK: {len(items)} itens | {len(models)} modelos | {len(analysts)} analistas | {len(families)} famílias | {len(obtention_types)} tipos de obtenção | {open_requests} solicitações abertas | {OUTPUT}')
