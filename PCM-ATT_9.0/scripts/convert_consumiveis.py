"""Converte a aba consumiveis do Excel para JSON usado pelo dashboard."""
from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(sys.argv[1]) if len(sys.argv) > 1 else Path('/home/ubuntu/upload/consumiveisPronta12.08.xlsx')
OUTPUT = PROJECT_ROOT / 'data' / 'consumiveis.json'


def text(value):
    return '' if value is None else str(value).strip()


def number(value):
    if isinstance(value, (int, float)):
        return round(float(value), 6)
    raw = text(value).replace('.', '').replace(',', '.')
    try:
        return round(float(raw), 6) if raw else 0
    except ValueError:
        return 0


def compact(value):
    value = number(value)
    return int(value) if value.is_integer() else value


def serialise(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


wb = load_workbook(SOURCE, read_only=True, data_only=True)
if 'consumiveis' not in wb.sheetnames:
    raise RuntimeError('A aba consumiveis não foi encontrada na planilha.')

ws = wb['consumiveis']
header_row = 3
headers = [text(value) for value in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
if len(headers) < 14 or headers[0] != 'Código':
    raise RuntimeError('A aba consumiveis não possui o cabeçalho esperado na linha 3.')

order_columns = []
for index, header in enumerate(headers):
    match = re.search(r'PED\s+(\d{2}/\d{4})', header.upper())
    if match:
        order_columns.append((index, f'PED {match.group(1)}'))

items = []
seen = set()
for values in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
    code = text(values[0] if len(values) > 0 else '')
    if not code or code in seen or code.lower() == 'código':
        continue
    seen.add(code)
    orders = {label: compact(values[index]) for index, label in order_columns if index < len(values)}
    items.append({
        'code': code,
        'description': text(values[1] if len(values) > 1 else ''),
        'stock': compact(values[3] if len(values) > 3 else 0),
        'purchaseQty': compact(values[4] if len(values) > 4 else 0),
        'minStock': compact(values[5] if len(values) > 5 else 0),
        'maxStock': compact(values[6] if len(values) > 6 else 0),
        'stockStatus': text(values[12] if len(values) > 12 else ''),
        'maxStatus': text(values[13] if len(values) > 13 else ''),
        'purchaseDate': serialise(values[14] if len(values) > 14 else None),
        'nextPurchaseDate': serialise(values[15] if len(values) > 15 else None),
        'reviewStatus': text(values[16] if len(values) > 16 else ''),
        'orders': orders,
        'orderTotal': compact(values[22] if len(values) > 22 else 0),
        'buyStatus': text(values[23] if len(values) > 23 else ''),
        'averageCost': compact(values[25] if len(values) > 25 else 0),
    })

payload = {
    'sourceFile': SOURCE.name,
    'generatedAt': date.today().isoformat(),
    'sheet': 'consumiveis',
    'columns': {'code': 'A', 'description': 'B', 'stock': 'D', 'purchaseQty': 'E', 'minStock': 'F', 'maxStock': 'G', 'stockStatus': 'M', 'maxStatus': 'N'},
    'months': [label for _, label in order_columns],
    'items': items,
}
OUTPUT.parent.mkdir(parents=True, exist_ok=True)
OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
print(f'OK: {len(items)} consumíveis convertidos para {OUTPUT}')
