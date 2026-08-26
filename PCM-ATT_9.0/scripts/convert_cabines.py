from pathlib import Path
from datetime import date
import json
import re
import sys

from openpyxl import load_workbook

HEADER_ROW = 2
MODEL_ROW = 1
FIRST_DATA_ROW = 3
OUTPUT_DEFAULT = Path(__file__).resolve().parents[1] / 'data' / 'cabines.json'


def text(value):
    return '' if value is None else str(value).strip()


def number(value):
    if isinstance(value, (int, float)):
        return float(value or 0)
    raw = text(value).replace('.', '').replace(',', '.')
    try:
        return float(raw) if raw else 0
    except ValueError:
        return 0


def compact(value):
    value = float(value or 0)
    return int(value) if value.is_integer() else round(value, 4)


def model_name(value, used):
    base = re.sub(r'^cabine\s*', '', text(value), flags=re.I).strip() or 'Modelo'
    base = re.sub(r'\s+', ' ', base)
    result = base.upper()
    suffix = 2
    while result in used:
        result = f'{base.upper()} ({suffix})'
        suffix += 1
    return result


def extrair(origem: Path, destino: Path) -> None:
    workbook = load_workbook(origem, read_only=False, data_only=True, keep_vba=True)
    sheet_name = next((name for name in ('cabines', 'Cabines') if name in workbook.sheetnames), None)
    if not sheet_name:
        raise ValueError('A aba cabines não foi encontrada na planilha.')
    ws = workbook[sheet_name]

    blocks = []
    used_models = set()
    for col in range(1, ws.max_column + 1):
        header = text(ws.cell(HEADER_ROW, col).value).lower()
        if header not in {'código', 'codigo', 'código item', 'codigo item'}:
            continue
        model = ''
        for candidate_col in range(max(1, col), min(ws.max_column, col + 8) + 1):
            candidate = text(ws.cell(MODEL_ROW, candidate_col).value)
            if candidate:
                model = model_name(candidate, used_models)
                break
        if not model:
            model = model_name(f'Modelo {len(blocks) + 1}', used_models)
        used_models.add(model)
        blocks.append({
            'model': model,
            'code': col,
            'description': col + 1,
            'unit': None,
            'stock': col + 3,
            'need': col + 2,
        })

    items = {}
    model_order = [block['model'] for block in blocks]
    for block in blocks:
        for row in range(FIRST_DATA_ROW, ws.max_row + 1):
            code = text(ws.cell(row, block['code']).value)
            if not code or code.lower() in {'código', 'codigo', 'código item', 'codigo item'}:
                continue
            description = text(ws.cell(row, block['description']).value)
            item = items.setdefault(code, {
                'code': code,
                'description': description,
                'unit': 'UN',
                'stock': 0,
                'modelNeeds': {model: 0 for model in model_order},
                'lastMovement': 'não tem',
            })
            if not item['description'] and description:
                item['description'] = description
            stock = compact(number(ws.cell(row, block['stock']).value))
            need = compact(number(ws.cell(row, block['need']).value))
            if item['stock'] == 0 and stock:
                item['stock'] = stock
            item['modelNeeds'][block['model']] = compact(number(item['modelNeeds'].get(block['model'], 0)) + need)

    output_items = []
    for item in items.values():
        item['totalNeed'] = compact(sum(number(value) for value in item['modelNeeds'].values()))
        item['modelCount'] = sum(1 for value in item['modelNeeds'].values() if number(value) > 0)
        output_items.append(item)
    output_items.sort(key=lambda item: (-number(item['totalNeed']), item['code']))

    payload = {
        'sourceFile': origem.name,
        'sourceSheet': sheet_name,
        'generatedAt': date.today().isoformat(),
        'models': model_order,
        'items': output_items,
        'meta': {
            'headerRow': HEADER_ROW,
            'modelRow': MODEL_ROW,
            'firstDataRow': FIRST_DATA_ROW,
            'codeBlocks': len(blocks),
        },
    }
    destino.parent.mkdir(parents=True, exist_ok=True)
    destino.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'CABINES_OK {len(output_items)} itens | {len(model_order)} modelos | {destino}')


if __name__ == '__main__':
    if len(sys.argv) not in {2, 3}:
        raise SystemExit('Uso: python convert_cabines.py origem.xlsm [destino.json]')
    extrair(Path(sys.argv[1]), Path(sys.argv[2]) if len(sys.argv) == 3 else OUTPUT_DEFAULT)
