from openpyxl import load_workbook
from pathlib import Path
source = Path('/home/ubuntu/upload/CópiadeEXPLOSAO_04.08.xlsm')
wb = load_workbook(source, read_only=True, data_only=True, keep_vba=True)
for ws in wb.worksheets:
    print(ws.title, ws.max_row, ws.max_column)
