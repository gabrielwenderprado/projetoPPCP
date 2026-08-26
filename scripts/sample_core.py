from openpyxl import load_workbook
source='/home/ubuntu/upload/CópiadeEXPLOSAO_04.08.xlsm'
wb=load_workbook(source, read_only=True, data_only=True, keep_vba=True)
for name, max_col in [('Estoque', 12), ('Programacao', 45), ('Obtencao', 20), ('estoque segurança', 10)]:
    ws=wb[name]
    print(f'### {name}')
    for row in ws.iter_rows(min_row=1, max_row=6, max_col=max_col, values_only=True):
        print([str(v)[:90] if v is not None else '' for v in row])
